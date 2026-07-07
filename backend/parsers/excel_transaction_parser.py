from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def parse_excel_rows(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if not any(headers):
        raise ValueError("Excel file is missing a header row.")

    parsed_rows: list[dict[str, object]] = []
    for values in rows[1:]:
        parsed_rows.append({
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
        })
    return parsed_rows
